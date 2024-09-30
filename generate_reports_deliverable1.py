import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.worksheet.table import Table, TableStyleInfo

def create_excel_report(df):
    path_to_save = 'Financial_Analysis.xlsx'
    
    with pd.ExcelWriter(path_to_save, engine='openpyxl') as writer:
        # Total transactions per client, per currency
        pivot_df = df.pivot_table(index=['Client', 'Currency'], values='Transaction', aggfunc='count')
        pivot_df.to_excel(writer, sheet_name='Transactions Summary')
        
        # Totals by client in USD
        totals_usd = df.groupby('Client')['Transaction'].sum()
        totals_usd.to_excel(writer, sheet_name='Total by Client USD')

        workbook = writer.book
        sheet1 = workbook['Transactions Summary']
        sheet2 = workbook['Total by Client USD']
        
        # Set AutoFilter on the DataFrame ranges
        sheet1.auto_filter.ref = sheet1.dimensions
        sheet2.auto_filter.ref = sheet2.dimensions
        
        # Create a bar chart for the top 10 clients in "Transactions Summary"
        top_clients = pivot_df.groupby('Client')['Transaction'].sum().nlargest(10).reset_index()
        
        # Add top 10 clients data to the sheet
        start_row = sheet1.max_row + 2
        for r_idx, row in enumerate(dataframe_to_rows(top_clients, index=False, header=True), start_row):
            for c_idx, value in enumerate(row, 1):
                sheet1.cell(row=r_idx, column=c_idx, value=value)

        # Create a chart for top 10 clients
        if len(top_clients) > 0:
            chart1 = BarChart()
            data1 = Reference(sheet1, min_col=2, min_row=start_row + 1, max_row=start_row + len(top_clients))
            categories1 = Reference(sheet1, min_col=1, min_row=start_row + 2, max_row=start_row + len(top_clients))
            chart1.add_data(data1, titles_from_data=True)
            chart1.set_categories(categories1)
            chart1.title = "Top 10 Clients by Transactions"
            chart1.x_axis.title = "Client"
            chart1.y_axis.title = "Transaction Count"
            
            # Add data labels
            chart1.dLbls = DataLabelList()
            chart1.dLbls.showVal = True
            
            # Place chart at a valid position
            sheet1.add_chart(chart1, "E4")
        
        # Create a pie chart for the top 5 currencies
        top_currencies = df['Currency'].value_counts().nlargest(5).reset_index()
        top_currencies.columns = ['Currency', 'Count']
        
        start_row_currency = sheet1.max_row + 2
        for r_idx, row in enumerate(dataframe_to_rows(top_currencies, index=False, header=True), start_row_currency):
            for c_idx, value in enumerate(row, 1):
                sheet1.cell(row=r_idx, column=c_idx, value=value)

        if len(top_currencies) > 0:
            pie_chart = PieChart()
            pie_data = Reference(sheet1, min_col=2, min_row=start_row_currency + 1, max_row=start_row_currency + len(top_currencies))
            pie_categories = Reference(sheet1, min_col=1, min_row=start_row_currency + 2, max_row=start_row_currency + len(top_currencies))
            pie_chart.add_data(pie_data, titles_from_data=True)
            pie_chart.set_categories(pie_categories)
            pie_chart.title = "Top 5 Currencies Distribution"
            
            # Add data labels
            pie_chart.dLbls = DataLabelList()
            pie_chart.dLbls.showVal = True
            
            # Place pie chart below the bar chart
            sheet1.add_chart(pie_chart, "E16")
        
        # Create a bar chart for the top 10 clients in "Total by Client USD"
        top_10_totals_usd = totals_usd.nlargest(10).reset_index()
        
        # Add top 10 clients data to the sheet
        start_row_usd = sheet2.max_row + 2
        for r_idx, row in enumerate(dataframe_to_rows(top_10_totals_usd, index=False, header=True), start_row_usd):
            for c_idx, value in enumerate(row, 1):
                sheet2.cell(row=r_idx, column=c_idx, value=value)
        
        # Create chart for top 10 clients in USD
        if len(top_10_totals_usd) > 0:
            chart2 = BarChart()
            values = Reference(sheet2, min_col=2, min_row=start_row_usd + 1, max_row=start_row_usd + len(top_10_totals_usd))
            categories = Reference(sheet2, min_col=1, min_row=start_row_usd + 2, max_row=start_row_usd + len(top_10_totals_usd))
            chart2.add_data(values, titles_from_data=True)
            chart2.set_categories(categories)
            chart2.title = "Total USD per Top 10 Clients"
            chart2.y_axis.title = "Total in USD"
            chart2.x_axis.title = "Client"
            
            # Add data labels
            chart2.dLbls = DataLabelList()
            chart2.dLbls.showVal = True
            
            # Place chart at a valid position
            sheet2.add_chart(chart2, "E4")
        
        # Add formatting and a table style to the pivot table sheet
        tab = Table(displayName="TransactionTable", ref="A1:C{}".format(pivot_df.shape[0]+1))
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        sheet1.add_table(tab)
