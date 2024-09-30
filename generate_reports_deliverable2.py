import pandas as pd
import requests
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from datetime import datetime

def get_countries_by_subregion(subregion):
    """
    Fetches the list of countries for a given subregion using the REST Countries API.
    Converts country names to uppercase.
    """
    url = f'https://restcountries.com/v3.1/subregion/{subregion}'
    response = requests.get(url)
    
    if response.status_code == 200:
        countries = response.json()
        country_names = [country['name']['common'].upper() for country in countries]
        return country_names
    else:
        print(f"Error fetching data for {subregion}: {response.status_code}")
        return []

def get_latin_american_countries():
    """
    Combines countries from the specified subregions (South America, Central America, Caribbean).
    Returns a list of all unique country names in uppercase.
    """
    subregions = ['South America', 'Central America', 'Caribbean']
    all_countries = set()  # Use a set to avoid duplicates
    
    for subregion in subregions:
        countries = get_countries_by_subregion(subregion)
        all_countries.update(countries)  # Add countries to the set
    
    return list(all_countries)

def get_last_quarter_dates():
    """
    Returns the start and end dates of the last quarter of the year 2024.
    """
    max_year = 2024
    start_date = pd.Timestamp(max_year, 10, 1)
    end_date = pd.Timestamp(max_year, 12, 31)
    
    # Convert to datetime.date to match DataFrame's date type
    return start_date.date(), end_date.date()

def create_deliverable_2(df):
    """
    Creates an Excel file with two sheets:
    1. 'LATAM Transactions': Contains transactions for Latin American countries.
    2. 'Specified Countries Q4 2024': Contains transactions with Russia, Cuba, China, and Venezuela for Q4 2024.
    
    Additionally, generates charts for data analysis.
    """
    path_to_save = 'Latin_America_Transactions.xlsx'
    
    # Convert the 'Date' column to datetime.date type
    df['Date'] = pd.to_datetime(df['Date']).dt.date
    
    # Filter transactions for Latin American countries
    latin_american_countries = get_latin_american_countries()
    df_latam = df[df['Country'].isin(latin_american_countries)]

    # Filter transactions with specific countries (Russia, Cuba, China, Venezuela) for Q4 2024
    specific_countries = ['RUSSIA', 'CUBA', 'CHINA', 'VENEZUELA']
    start_date, end_date = get_last_quarter_dates()
    df_specific = df[(df['Country'].isin(specific_countries)) & 
                     (df['Date'] >= start_date) & 
                     (df['Date'] <= end_date)]

    # Check if there are transactions for the specific countries
    if df_specific.empty:
        print("No transactions found for Russia, Cuba, China, or Venezuela in Q4 2024.")
        return
    
    # Define columns to include in each sheet
    cols_basic = ['Client', 'Country', 'Currency', 'Transaction']
    cols_with_date = cols_basic + ['Date']
    
    with pd.ExcelWriter(path_to_save, engine='openpyxl') as writer:
        # Write 'LATAM Transactions' sheet
        df_latam.loc[:, cols_basic].to_excel(writer, sheet_name='LATAM Transactions', index=False)
        
        # Write 'Specified Countries Q4 2024' sheet
        df_specific.loc[:, cols_with_date].to_excel(writer, sheet_name='Specified Countries Q4 2024', index=False)
        
        # Get workbook and sheets
        workbook = writer.book
        sheet1 = workbook['LATAM Transactions']
        sheet2 = workbook['Specified Countries Q4 2024']
        
        # Apply filters to the sheets
        sheet1.auto_filter.ref = sheet1.dimensions
        sheet2.auto_filter.ref = sheet2.dimensions
        
        # Add charts
        # 1. Chart for 'LATAM Transactions' (Transactions by country)
        country_counts = df_latam['Country'].value_counts().reset_index()
        country_counts.columns = ['Country', 'Transaction Count']
        
        # Add data at the end of the sheet
        start_row_latam = sheet1.max_row + 2
        for r_idx, row in enumerate(dataframe_to_rows(country_counts, index=False, header=True), start_row_latam):
            for c_idx, value in enumerate(row, 1):
                sheet1.cell(row=r_idx, column=c_idx, value=value)
        
        # Create a bar chart for Latin American countries
        chart1 = BarChart()
        data1 = Reference(sheet1, min_col=2, min_row=start_row_latam + 1, max_row=start_row_latam + len(country_counts), max_col=2)
        categories1 = Reference(sheet1, min_col=1, min_row=start_row_latam + 2, max_row=start_row_latam + len(country_counts))
        chart1.add_data(data1, titles_from_data=True)
        chart1.set_categories(categories1)
        chart1.title = "Transactions by Country (LATAM)"
        chart1.x_axis.title = "Country"
        chart1.y_axis.title = "Transaction Count"
        # Place the chart in the top-right corner
        sheet1.add_chart(chart1, "E2")

        # 2. Chart for 'Specified Countries Q4 2024' (Group by country)
        country_counts_specific = df_specific['Country'].value_counts().reset_index()
        country_counts_specific.columns = ['Country', 'Transaction Count']
        
        # Add data at the end of the sheet
        start_row_specified = sheet2.max_row + 2
        for r_idx, row in enumerate(dataframe_to_rows(country_counts_specific, index=False, header=True), start_row_specified):
            for c_idx, value in enumerate(row, 1):
                sheet2.cell(row=r_idx, column=c_idx, value=value)
        
        # Create a bar chart for specified countries
        chart2 = BarChart()
        data2 = Reference(sheet2, min_col=2, min_row=start_row_specified + 1, max_row=start_row_specified + len(country_counts_specific), max_col=2)
        categories2 = Reference(sheet2, min_col=1, min_row=start_row_specified + 2, max_row=start_row_specified + len(country_counts_specific))
        chart2.add_data(data2, titles_from_data=True)
        chart2.set_categories(categories2)
        chart2.title = "Transactions by Country (Specified Q4 2024)"
        chart2.x_axis.title = "Country"
        chart2.y_axis.title = "Transaction Count"
        # Place the chart in the top-right corner
        sheet2.add_chart(chart2, "E2")
        
        # Create a pie chart to show the proportion of transactions by country
        pie_chart = PieChart()
        pie_data = Reference(sheet2, min_col=2, min_row=start_row_specified + 1, max_row=start_row_specified + len(country_counts_specific))
        pie_categories = Reference(sheet2, min_col=1, min_row=start_row_specified + 2, max_row=start_row_specified + len(country_counts_specific))
        pie_chart.add_data(pie_data, titles_from_data=True)
        pie_chart.set_categories(pie_categories)
        pie_chart.title = "Transaction Distribution by Country"
        # Place the pie chart to the right of the bar chart
        sheet2.add_chart(pie_chart, "E16")
        
    workbook.save(filename=path_to_save)
